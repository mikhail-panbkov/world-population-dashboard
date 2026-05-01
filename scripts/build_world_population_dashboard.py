from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "data" / "world_population_data.csv"
CLEANED_PATH = ROOT / "data" / "world_population_cleaned.csv"
OUTPUT_DIR = ROOT / "outputs" / "project3_world_population"
OUTPUT_PATH = OUTPUT_DIR / "world_population_dashboard.xlsx"

YEARS = [1970, 1980, 1990, 2000, 2010, 2015, 2020, 2022, 2023]
POPULATION_COLUMNS = [f"{year}_population" for year in YEARS]


def normalize_column_name(name: str) -> str:
    normalized = name.strip().lower().replace(" ", "_")
    return re.sub(r"_\(km.*\)", "", normalized)


def load_and_clean_data() -> pd.DataFrame:
    df = pd.read_csv(SOURCE_PATH)
    df.columns = [normalize_column_name(column) for column in df.columns]

    for column in ["growth_rate", "world_percentage"]:
        df[column] = (
            df[column]
            .astype(str)
            .str.replace("%", "", regex=False)
            .replace("nan", pd.NA)
            .astype(float)
        )

    numeric_columns = ["rank", "area", "density", *POPULATION_COLUMNS]
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df["change_since_1970"] = df["2023_population"] - df["1970_population"]
    df["growth_since_1970"] = df["2023_population"] / df["1970_population"]
    return df


def append_dataframe(sheet, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 28)

    sheet.freeze_panes = "A2"


def add_title(sheet, title: str, last_column: str) -> None:
    sheet.insert_rows(1)
    sheet.merge_cells(f"A1:{last_column}1")
    cell = sheet["A1"]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor="17324D")
    cell.font = Font(bold=True, color="FFFFFF", size=16)
    cell.alignment = Alignment(horizontal="left")
    sheet.row_dimensions[1].height = 26


def build_workbook(df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    continent_sheet = wb.create_sheet("Continent Summary")
    ranking_sheet = wb.create_sheet("Top Countries")
    raw_sheet = wb.create_sheet("Raw Data")
    notes_sheet = wb.create_sheet("Notes")

    trend = pd.DataFrame(
        {
            "year": YEARS,
            "world_population": [df[column].sum() for column in POPULATION_COLUMNS],
        }
    )

    population_1970 = df.groupby("continent")["1970_population"].sum()
    continent_summary = (
        df.groupby("continent", as_index=False)
        .agg(
            countries=("country", "count"),
            population_2023=("2023_population", "sum"),
            world_share=("world_percentage", "sum"),
            avg_growth_rate=("growth_rate", "mean"),
            median_density=("density", "median"),
            area=("area", "sum"),
            change_since_1970=("change_since_1970", "sum"),
        )
        .sort_values("population_2023", ascending=False)
    )
    continent_summary["growth_since_1970"] = (
        continent_summary["population_2023"]
        / population_1970.reindex(continent_summary["continent"]).to_numpy()
    )

    top_tables = [
        (
            "Top 10 by 2023 Population",
            df.nlargest(10, "2023_population")[
                ["country", "continent", "2023_population", "world_percentage", "growth_rate"]
            ],
        ),
        (
            "Top 10 by Growth Rate",
            df.nlargest(10, "growth_rate")[
                ["country", "continent", "growth_rate", "2023_population", "density"]
            ],
        ),
        (
            "Top 10 by Density",
            df.nlargest(10, "density")[["country", "continent", "density", "area", "2023_population"]],
        ),
        (
            "Top 10 by Absolute Change Since 1970",
            df.nlargest(10, "change_since_1970")[
                ["country", "continent", "1970_population", "2023_population", "change_since_1970", "growth_since_1970"]
            ],
        ),
    ]

    dashboard.append(["Metric", "Value"])
    dashboard.append(["2023 world population", int(df["2023_population"].sum())])
    dashboard.append(["Countries / territories", len(df)])
    dashboard.append(["Largest country", df.loc[df["2023_population"].idxmax(), "country"]])
    dashboard.append(["Largest continent", continent_summary.iloc[0]["continent"]])
    dashboard.append(["Average growth rate", round(df["growth_rate"].mean(), 2)])
    dashboard.append(["Population multiple vs 1970", round(df["2023_population"].sum() / df["1970_population"].sum(), 2)])
    dashboard.append([])
    dashboard.append(["Year", "World population"])
    for row in trend.itertuples(index=False):
        dashboard.append(list(row))
    dashboard.append([])
    dashboard.append(["Continent", "2023 population"])
    for row in continent_summary[["continent", "population_2023"]].itertuples(index=False):
        dashboard.append(list(row))
    style_sheet(dashboard)
    add_title(dashboard, "World Population Dashboard", "B")

    line_chart = LineChart()
    line_chart.title = "World Population Trend"
    line_chart.y_axis.title = "Population"
    line_chart.x_axis.title = "Year"
    line_chart.add_data(Reference(dashboard, min_col=2, min_row=10, max_row=19), titles_from_data=True)
    line_chart.set_categories(Reference(dashboard, min_col=1, min_row=11, max_row=19))
    dashboard.add_chart(line_chart, "D3")

    bar_chart = BarChart()
    bar_chart.title = "2023 Population by Continent"
    bar_chart.y_axis.title = "Population"
    bar_chart.add_data(Reference(dashboard, min_col=2, min_row=21, max_row=27), titles_from_data=True)
    bar_chart.set_categories(Reference(dashboard, min_col=1, min_row=22, max_row=27))
    dashboard.add_chart(bar_chart, "D20")

    append_dataframe(continent_sheet, continent_summary)
    style_sheet(continent_sheet)
    add_title(continent_sheet, "Continent Summary", "H")

    for title, table in top_tables:
        ranking_sheet.append([title])
        ranking_sheet.cell(ranking_sheet.max_row, 1).font = Font(bold=True)
        for row in dataframe_to_rows(table, index=False, header=True):
            ranking_sheet.append(row)
        ranking_sheet.append([])
    style_sheet(ranking_sheet)

    append_dataframe(raw_sheet, df)
    style_sheet(raw_sheet)

    for row in [
        ["Finding", "How it is shown"],
        ["Asia has the largest 2023 population.", "Continent summary and dashboard chart."],
        ["India and China are the two largest countries.", "Top population ranking."],
        ["Growth rate and total population tell different stories.", "Separate rankings for population, growth, and density."],
        ["The world population is much higher than in 1970.", "Trend chart and population multiple KPI."],
    ]:
        notes_sheet.append(row)
    style_sheet(notes_sheet)
    add_title(notes_sheet, "Analysis Notes", "B")

    return wb


def main() -> None:
    df = load_and_clean_data()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(CLEANED_PATH, index=False)
    build_workbook(df).save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
